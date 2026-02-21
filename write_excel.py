import openpyxl
import logging
import os
from datetime import datetime

# 配置日志
log_dir = 'logs'
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, f'digikey_{datetime.now().strftime("%Y%m%d")}.log')

# 创建日志记录器
logger = logging.getLogger('excel_handler')
logger.setLevel(logging.INFO)

# 创建文件处理器
file_handler = logging.FileHandler(log_file, encoding='utf-8')
file_handler.setLevel(logging.INFO)

# 创建控制台处理器
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)

# 创建格式化器
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

# 添加处理器到日志记录器
logger.addHandler(file_handler)
logger.addHandler(console_handler)

def read_excel_data(excel_path, sheet_name, header_name, max_search_rows=10, return_header_info=False):
    """
    读取Excel文件中指定表头列的数据
    
    参数:
        excel_path: Excel文件路径
        sheet_name: 工作表名称
        header_name: 表头名称
        max_search_rows: 最大搜索行数，用于查找表头
        return_header_info: 是否返回表头信息
    """
    logger.info(f"开始读取Excel文件: {excel_path}, 工作表: {sheet_name}, 表头: {header_name}")
    
    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(excel_path)
        logger.info(f"成功加载Excel文件: {excel_path}")
        
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            error_msg = f"工作表 '{sheet_name}' 不存在"
            logger.error(error_msg)
            raise Exception(error_msg)
            
        # 选择工作表
        sheet = workbook[sheet_name]
        logger.info(f"成功选择工作表: {sheet_name}")
        
        # 智能查找表头所在行和列
        header_row = None
        header_column = None
        
        # 搜索前max_search_rows行查找表头
        for row_num in range(1, min(max_search_rows + 1, sheet.max_row + 1)):
            for cell in sheet[row_num]:
                if cell.value == header_name:
                    header_row = row_num
                    header_column = cell.column
                    logger.info(f"找到表头 '{header_name}' 在第 {header_row} 行第 {header_column} 列")
                    break
            if header_row:
                break
                
        if not header_row or not header_column:
            error_msg = f"表头 '{header_name}' 未找到（搜索了前{max_search_rows}行）"
            logger.error(error_msg)
            raise Exception(error_msg)
        
        # 从表头下方读取所有数据
        data = []
        start_row = header_row + 1  # 从表头下一行开始读取
        for row in sheet.iter_rows(min_row=start_row, min_col=header_column, max_col=header_column):
            if row[0].value:
                data.append(str(row[0].value).strip())
        
        if not data:
            logger.warning("未找到有效数据")
            if return_header_info:
                return [], header_row, header_column
            return []
        
        logger.info(f"成功读取 {len(data)} 条数据")
        
        if return_header_info:
            return data, header_row, header_column
        return data
        
    except Exception as e:
        error_msg = f"读取Excel文件时发生错误: {str(e)}"
        logger.error(error_msg, exc_info=True)
        raise Exception(error_msg)


def write_excel_data(excel_path, sheet_name, header_name, data, max_search_rows=10, header_row=None):
    """
    写入数据到Excel文件中指定表头列
    
    参数:
        excel_path: Excel文件路径
        sheet_name: 工作表名称
        header_name: 表头名称
        data: 要写入的数据列表
        max_search_rows: 最大搜索行数，用于查找表头
        header_row: 表头行号（如果已知，可提高性能）
    """
    logger.info(f"开始写入Excel文件: {excel_path}, 工作表: {sheet_name}, 表头: {header_name}, 数据量: {len(data) if data else 0}")
    
    try:
        if not data:
            error_msg = '数据列表不能为空'
            logger.error(error_msg)
            return {'status': 'error', 'message': error_msg}
            
        workbook = openpyxl.load_workbook(excel_path)
        logger.info(f"成功加载Excel文件: {excel_path}")
        
        if sheet_name not in workbook.sheetnames:
            error_msg = f"工作表 '{sheet_name}' 不存在"
            logger.error(error_msg)
            return {'status': 'error', 'message': error_msg}
            
        sheet = workbook[sheet_name]
        logger.info(f"成功选择工作表: {sheet_name}")
        
        # 查找表头所在列（如果已知表头行号，直接在该行查找）
        header_column = None
        
        if header_row is not None:
            # 使用已知的表头行号
            logger.info(f"使用已知表头行号: {header_row}")
            for cell in sheet[header_row]:
                if cell.value == header_name:
                    header_column = cell.column
                    logger.info(f"找到现有表头 '{header_name}' 在第 {header_row} 行第 {header_column} 列")
                    break
        else:
            # 搜索前max_search_rows行查找表头
            for row_num in range(1, min(max_search_rows + 1, sheet.max_row + 1)):
                for cell in sheet[row_num]:
                    if cell.value == header_name:
                        header_row = row_num
                        header_column = cell.column
                        logger.info(f"找到现有表头 '{header_name}' 在第 {header_row} 行第 {header_column} 列")
                        break
                if header_row:
                    break
        
        if not header_column:
            # 如果未找到表头，创建新表头
            if header_row is None:
                header_row = 1  # 默认在第一行创建
            header_column = sheet.max_column + 1
            sheet.cell(row=header_row, column=header_column, value=header_name)
            logger.info(f"创建新表头 '{header_name}' 在第 {header_row} 行第 {header_column} 列")
        
        # 直接使用表头行号确定起始行
        start_row = header_row + 1
        
        # 写入数据，确保行数匹配
        for i, value in enumerate(data):
            row_num = start_row + i
            sheet.cell(row=row_num, column=header_column, value=value)
        
        workbook.save(excel_path)
        logger.info(f"成功保存Excel文件: {excel_path}")
        
        success_msg = f"成功写入 {len(data)} 条数据到Excel文件"
        logger.info(success_msg)
        return {'status': 'success', 'message': success_msg}
        
    except Exception as e:
        error_msg = f"写入Excel文件时发生错误: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {'status': 'error', 'message': error_msg}

def write_multiple_columns(excel_path, sheet_name, columns_data, max_search_rows=10, reference_header=None, reference_header_row=None):
    """
    写入多列数据到Excel文件
    
    参数:
        excel_path: Excel文件路径
        sheet_name: 工作表名称
        columns_data: 字典，键为列名，值为数据列表
        max_search_rows: 最大搜索行数，用于查找表头
        reference_header: 参考表头名称（用于确定表头行号）
        reference_header_row: 参考表头行号（如果已知）
    """
    logger.info(f"开始写入多列数据到Excel文件: {excel_path}, 工作表: {sheet_name}, 列数: {len(columns_data) if columns_data else 0}")
    
    try:
        if not columns_data:
            error_msg = '列数据字典不能为空'
            logger.error(error_msg)
            return {'status': 'error', 'message': error_msg}
            
        workbook = openpyxl.load_workbook(excel_path)
        logger.info(f"成功加载Excel文件: {excel_path}")
        
        if sheet_name not in workbook.sheetnames:
            error_msg = f"工作表 '{sheet_name}' 不存在"
            logger.error(error_msg)
            return {'status': 'error', 'message': error_msg}
            
        sheet = workbook[sheet_name]
        logger.info(f"成功选择工作表: {sheet_name}")
        
        # 确定表头行号
        header_row = reference_header_row
        
        if header_row is None and reference_header:
            # 通过参考表头确定表头行号
            logger.info(f"通过参考表头 '{reference_header}' 确定表头行号")
            for row_num in range(1, min(max_search_rows + 1, sheet.max_row + 1)):
                for cell in sheet[row_num]:
                    if cell.value == reference_header:
                        header_row = row_num
                        logger.info(f"找到参考表头 '{reference_header}' 在第 {header_row} 行")
                        break
                if header_row:
                    break
        
        if header_row is None:
            # 如果没有参考表头，使用第一行作为默认表头行
            header_row = 1
            logger.info(f"使用默认表头行号: {header_row}")
        
        # 处理每一列数据
        for header_name, data in columns_data.items():
            logger.info(f"处理列 '{header_name}', 数据量: {len(data) if data else 0}")
            
            # 查找表头所在列
            header_column = None
            
            # 在已知的表头行中查找表头
            for cell in sheet[header_row]:
                if cell.value == header_name:
                    header_column = cell.column
                    logger.info(f"找到现有表头 '{header_name}' 在第 {header_row} 行第 {header_column} 列")
                    break
            
            if not header_column:
                # 如果未找到表头，创建新表头
                header_column = sheet.max_column + 1
                sheet.cell(row=header_row, column=header_column, value=header_name)
                logger.info(f"创建新表头 '{header_name}' 在第 {header_row} 行第 {header_column} 列")
            
            # 直接使用表头行号确定起始行
            start_row = header_row + 1
            
            # 写入数据，确保行数匹配
            for i, value in enumerate(data):
                row_num = start_row + i
                sheet.cell(row=row_num, column=header_column, value=value)
        
        workbook.save(excel_path)
        logger.info(f"成功保存Excel文件: {excel_path}")
        
        success_msg = f"成功写入 {len(columns_data)} 列数据到Excel文件"
        logger.info(success_msg)
        return {'status': 'success', 'message': success_msg}
        
    except Exception as e:
        error_msg = f"写入多列数据到Excel文件时发生错误: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {'status': 'error', 'message': error_msg}

if __name__ == '__main__':
    # 示例用法 - 读取数据
    excel_path = input("请输入Excel文件路径: ")
    sheet_name = input("请输入工作表名称: ")
    header_name = input("请输入表头名称: ")
    mode = input("请输入操作模式 (读取/写入/多列写入): ")
    if mode == '读取':
        result = read_excel_data(excel_path, sheet_name, header_name)
        print(f"读取状态: {result['status']}")
        print(f"读取消息: {result['message']}")
        if 'data' in result and result['data']:
            print("数据列表:")
            for item in result['data']:
                print(f"- {item}")
    elif mode == '写入':
        data = input("请输入要写入的数据 (逗号分隔): ").split(',')
        data = [item.strip() for item in data]
        result = write_excel_data(excel_path, sheet_name, header_name, data)
        print(f"写入状态: {result['status']}")
        print(f"写入消息: {result['message']}")
    elif mode == '多列写入':
        columns_data = {}
        while True:
            col_name = input("请输入列名 (留空结束): ")
            if not col_name:
                break
            col_data = input(f"请输入列 '{col_name}' 的数据 (逗号分隔): ").split(',')
            col_data = [item.strip() for item in col_data]
            columns_data[col_name] = col_data
        
        result = write_multiple_columns(excel_path, sheet_name, columns_data)
        print(f"多列写入状态: {result['status']}")
        print(f"多列写入消息: {result['message']}")
    else:
        print("无效的操作模式")


    
    
        
