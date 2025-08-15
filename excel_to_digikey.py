import openpyxl
#111
import json
import os
import sys
from digikey import digikey_api

def process_excel_to_digikey(excel_path, sheet_name, header_name):
    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(excel_path)
        if sheet_name not in workbook.sheetnames:
            return {'status': 'error', 'message': f"工作表'{sheet_name}'不存在"}
            
        sheet = workbook[sheet_name]
        
        # 查找表头所在列
        header_column = None
        for cell in sheet[1]:
            if cell.value == header_name:
                header_column = cell.column
                break
                
        if not header_column:
            return {'status': 'error', 'message': f"表头'{header_name}'未找到"}
        
        # 收集零件编号
        part_numbers = []
        for row in sheet.iter_rows(min_row=2, min_col=header_column, max_col=header_column):
            if row[0].value:
                part_numbers.append(str(row[0].value).strip())
        
        if not part_numbers:
            return {'status': 'error', 'message': "未找到有效的零件编号"}
        
        # 查询DigiKey API
        results = {}
        total = len(part_numbers)
        print("\n开始查询DigiKey API...")
        
        for i, part in enumerate(part_numbers, 1):
            # 显示进度
            progress = i / total * 100
            sys.stdout.write(f"\r处理进度: {i}/{total} ({progress:.1f}%) - 当前处理: {part}")
            sys.stdout.flush()
            
            api_result = digikey_api(part)
            if api_result['status'] == 'success':
                results[part] = api_result['product_status']
            else:
                results[part] = f"查询失败: {api_result.get('message', '未知错误')}"
        
        print("\n处理完成！")
        
        # 保存结果到data.json
        data_file = os.path.join(os.path.dirname(__file__), 'data.json')
        with open(data_file, 'w') as f:
            json.dump(results, f, indent=2)
            
        # 将结果写回Excel文件
        output_header = input("请输入要写入的Excel表头名称: ")
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook[sheet_name]
        
        # 查找输出表头所在列
        output_column = None
        for cell in sheet[1]:
            if cell.value == output_header:
                output_column = cell.column
                break
                
        if not output_column:
            # 如果表头不存在，创建新列
            output_column = sheet.max_column + 1
            sheet.cell(row=1, column=output_column, value=output_header)
        
        # 写入状态数据
        row_index = 2
        for part in part_numbers:
            if part in results:
                sheet.cell(row=row_index, column=output_column, value=results[part])
            row_index += 1
            
        workbook.save(excel_path)
        
        return {'status': 'success', 'message': f"成功处理{len(results)}个零件并写回Excel", 'data': results}
        
    except Exception as e:
        return {'status': 'error', 'message': f"处理过程中发生错误: {str(e)}"}

if __name__ == "__main__":
    # 示例用法
    excel_path = input("请输入Excel文件路径: ")
    sheet_name = input("请输入工作表名称: ")
    header_name = input("请输入表头名称: ")
    
    result = process_excel_to_digikey(excel_path, sheet_name, header_name)
    print(f"处理结果: {result['status']}")
    print(f"消息: {result['message']}")
    if 'data' in result:
        print(f"查询结果已保存到data.json")